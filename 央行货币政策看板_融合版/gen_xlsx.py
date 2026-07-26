#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_xlsx.py — 由 pipeline_report.json 生成筛选明细工作簿"""
import json
import sys
import os
import zipfile
import shutil
import re

_CODE_DIR = os.path.dirname(os.path.abspath(__file__))
if _CODE_DIR not in sys.path:
    sys.path.insert(0, _CODE_DIR)

import config
import extractor
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

REPORT = config.REPORT_PATH
CACHE = config.CACHE_PATH
OUT = os.path.join(config.PROJECT_ROOT, '央行看板_筛选明细.xlsx')

with open(REPORT, encoding='utf-8') as f:
    report = json.load(f)
with open(CACHE, encoding='utf-8') as f:
    cache = json.load(f)
link_map = {(a.get('title', ''), a.get('date', '')): a.get('link', '') for a in cache}
# 文章页权威标题映射（剔除明细表也享受页面标题清洗）
title_map = {(a.get('title', ''), a.get('date', '')): a.get('title_page', '') for a in cache}


def show_title(k):
    """展示标题：report里的clean_title > 缓存title_page > 模式清洗兜底"""
    return (k.get('clean_title') or title_map.get((k['title'], k['date']))
            or extractor.clean_title(k['title']))
s = report['summary']
DIMS = ['经济形势判断', '工作重点', '货币政策取向', '资金态度', '债券市场态度']
header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")


def hstyle(c):
    c.fill = header_fill
    c.font = Font(color="FFFFFF", bold=True)


wb = Workbook()
ws = wb.active
ws.title = "封面"
ws.sheet_view.showGridLines = False
ws['B2'] = "央行货币政策跟踪看板 · 全量筛选明细"
ws['B2'].font = Font(size=18, bold=True)
ws.row_dimensions[2].height = 32
ws['B4'] = ("数据来源：金融时报/金融新闻网（financialnews.com.cn）马梅若、马玲署名文章，"
            "2024-08-01 至 2026-07-21｜筛选规则：2026-07-21 定稿版")
ws['B4'].font = Font(size=11, color="666666")
metrics = [("爬取文章总数（目标作者）", s['total_crawled']),
           ("看板保留", s['final_kept']),
           ("剔除合计", s['total_crawled'] - s['final_kept']),
           ("① 标题分类剔除", s['removed_by_stage'].get('①标题分类', 0)),
           ("② 排除规则剔除", s['removed_by_stage'].get('②排除规则', 0)),
           ("③ 观点提取剔除", s['removed_by_stage'].get('③观点提取', 0))]
ws['B6'] = "关键数据"
ws['B6'].font = Font(size=13, bold=True)
for j, h in enumerate(["项目", "数量（篇）"]):
    hstyle(ws.cell(row=7, column=2 + j, value=h))
for i, (k, v) in enumerate(metrics):
    ws.cell(row=8 + i, column=2, value=k)
    ws.cell(row=8 + i, column=3, value=v).font = Font(bold=True, color="0066CC")
ws['B16'] = "Sheet 索引"
ws['B16'].font = Font(size=13, bold=True)
for j, h in enumerate(["Sheet", "说明"]):
    hstyle(ws.cell(row=17, column=2 + j, value=h))
for i, (k, v) in enumerate([
        ("保留明细", "看板最终保留的 %d 篇文章及各维度摘句数" % s['final_kept']),
        ("剔除明细", "%d 篇被剔除文章的阶段与原因" % (s['total_crawled'] - s['final_kept']))]):
    ws.cell(row=18 + i, column=2, value=k)
    ws.cell(row=18 + i, column=3, value=v)
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 55
ws['B21'] = "注：保留/剔除明细中每行均附文章原始链接（financialnews.com.cn）作为来源；标题为清洗后版本。"
ws['B21'].font = Font(size=10, color="999999")

ws2 = wb.create_sheet("保留明细")
ws2.sheet_view.showGridLines = False
headers = ["日期", "标题", "作者", "标签"] + ["%s（条）" % d for d in DIMS] + ["来源链接"]
for j, h in enumerate(headers):
    hstyle(ws2.cell(row=2, column=2 + j, value=h))
ws2.row_dimensions[2].height = 22
for i, k in enumerate(sorted(report['kept'], key=lambda x: x['date'], reverse=True)):
    r = 3 + i
    d8 = k['date']
    ws2.cell(row=r, column=2, value="%s-%s-%s" % (d8[:4], d8[4:6], d8[6:8]))
    ws2.cell(row=r, column=3, value=show_title(k))
    ws2.cell(row=r, column=4, value=k['author'])
    ws2.cell(row=r, column=5, value=k['tag'])
    for j, dd in enumerate(DIMS):
        ws2.cell(row=r, column=6 + j, value=k['dims'].get(dd, 0))
    ws2.cell(row=r, column=11, value=link_map.get((k['title'], k['date']), ''))
for j, w in enumerate([11, 48, 8, 15, 13, 11, 14, 11, 14, 52]):
    ws2.column_dimensions[chr(ord('B') + j)].width = w
ws2.auto_filter.ref = "B2:K%d" % (2 + len(report['kept']))

ws3 = wb.create_sheet("剔除明细")
ws3.sheet_view.showGridLines = False
for j, h in enumerate(["日期", "标题", "作者", "标签", "剔除阶段", "剔除原因", "来源链接"]):
    hstyle(ws3.cell(row=2, column=2 + j, value=h))
ws3.row_dimensions[2].height = 22
removed_sorted = sorted(report['removed'], key=lambda x: x['date'], reverse=True)
for i, k in enumerate(removed_sorted):
    r = 3 + i
    d8 = k['date']
    ws3.cell(row=r, column=2, value="%s-%s-%s" % (d8[:4], d8[4:6], d8[6:8]))
    ws3.cell(row=r, column=3, value=show_title(k))
    ws3.cell(row=r, column=4, value=k.get('author', ''))
    ws3.cell(row=r, column=5, value=k.get('tag', ''))
    ws3.cell(row=r, column=6, value=k['stage'])
    ws3.cell(row=r, column=7, value=k['reason'])
    ws3.cell(row=r, column=8, value=link_map.get((k['title'], k['date']), ''))
for j, w in enumerate([11, 48, 8, 15, 12, 42, 52]):
    ws3.column_dimensions[chr(ord('B') + j)].width = w
ws3.auto_filter.ref = "B2:H%d" % (2 + len(removed_sorted))

wb.save(OUT)


def fix_rels_absolute_paths(xlsx_path):
    """openpyxl 3.1.5 会在 workbook.xml.rels 写入绝对路径导致校验失败，修正为相对路径"""
    tmp = xlsx_path + '.tmp'
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
            zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/_rels/workbook.xml.rels':
                data = re.sub(rb'Target="/xl/', b'Target="', data)
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)


fix_rels_absolute_paths(OUT)
print("saved", len(report['kept']), len(removed_sorted))
